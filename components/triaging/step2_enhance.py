# step2_enhance.py - UPDATED TO USE INTELLIGENT TEMPLATE GENERATOR
import streamlit as st
import os
import re
import time
import traceback
import hashlib
import pandas as pd
from io import BytesIO
from routes.src.virustotal_integration import VirusTotalChecker


def contains_ip_not_vip(text):
    """Check if text contains 'ip' but not as part of 'vip'"""
    if "ip" not in text:
        return False
    import re

    ip_patterns = [
        r"\bip\b",
        r"ip\s+address",
        r"ip\s+reputation",
        r"source\s+ip",
    ]
    return any(re.search(pattern, text, re.IGNORECASE) for pattern in ip_patterns)


def _save_step_data(step_num: int, rule_number: str, data_type: str):
    """Helper to save step data immediately"""
    if data_type == "remark":
        remark_key = f"remark_step_{step_num}_{rule_number}"
        input_key = f"remark_input_{step_num}"
        if input_key in st.session_state:
            st.session_state.step_remarks[remark_key] = st.session_state[input_key]

    elif data_type == "output":
        output_key = f"output_step_{step_num}_{rule_number}"
        input_key = f"output_input_{step_num}"
        if input_key in st.session_state:
            st.session_state.step_outputs[output_key] = st.session_state[input_key]


def _check_virustotal_auto(ip_address: str) -> dict:
    """Auto-check IP reputation using VirusTotal"""
    if "vt_checker" not in st.session_state:
        st.session_state.vt_checker = VirusTotalChecker()

    checker = st.session_state.vt_checker
    return checker.check_ip_reputation(ip_address.strip(), method="auto")


def _unlock_predictions(excel_data: bytes, filename: str, rule_number: str):
    """Callback to unlock predictions tab AND immediately upload file"""
    st.session_state.triaging_complete = True
    st.session_state.predictions_excel_data = excel_data
    st.session_state.predictions_excel_filename = filename
    st.session_state.predictions_rule_number = rule_number

    success = _upload_to_predictions_api(excel_data, filename)

    if success:
        st.session_state.predictions_uploaded = True
        st.session_state.show_predictions_unlock_message = True
        print(f"✅ Successfully uploaded {filename} to predictions API")
    else:
        st.session_state.predictions_uploaded = False
        st.session_state.show_predictions_unlock_message = True
        print(f"❌ Failed to upload {filename} to predictions API")


def _upload_to_predictions_api(excel_data: bytes, filename: str):
    """Upload Excel file to predictions API immediately"""
    try:
        import os
        from api_client.predictions_api_client import get_predictions_client

        final_api_key = os.getenv("GOOGLE_API_KEY")
        predictions_api_url = os.getenv(
            "PREDICTIONS_API_URL", "http://localhost:8000/predictions"
        )

        client = get_predictions_client(predictions_api_url, final_api_key)

        file_obj = BytesIO(excel_data)

        with st.spinner("📤 Uploading to predictions API..."):
            upload_result = client.upload_excel_bytes(file_obj, filename)

        if upload_result.get("success"):
            st.session_state.predictions_uploaded = True
            st.session_state.predictions_upload_result = upload_result
            st.session_state.predictions_file_data = excel_data
            st.session_state.predictions_filename = filename
            print(f"✅ Successfully uploaded {upload_result.get('total_rows', 0)} rows")
            return True
        else:
            st.session_state.predictions_upload_error = upload_result.get(
                "error", "Unknown error"
            )
            print(f"❌ Upload failed: {upload_result.get('error')}")
            return False

    except Exception as e:
        st.session_state.predictions_upload_error = str(e)
        print(f"❌ Upload exception: {str(e)}")
        return False


def _get_enhancement_cache_key(rule_number: str, template_path: str) -> str:
    """Create unique cache key for template enhancement"""
    return f"enhanced_template_{rule_number}_{hashlib.md5(template_path.encode()).hexdigest()}"


def _export_template_with_remarks_and_outputs(
    template_df, remarks_dict, outputs_dict, rule_number
):
    """Export template with remarks AND outputs columns added"""
    export_df = template_df.copy()

    remarks_list = []
    outputs_list = []
    step_counter = 1

    for idx, row in export_df.iterrows():
        if pd.isna(row["Step"]) or str(row["Step"]).strip() == "":
            remarks_list.append("")
            outputs_list.append("")
        else:
            remark_key = f"remark_step_{step_counter}_{rule_number}"
            remark = remarks_dict.get(remark_key, "")
            remarks_list.append(remark)

            output_key = f"output_step_{step_counter}_{rule_number}"
            output = outputs_dict.get(output_key, "")
            outputs_list.append(output)

            step_counter += 1

    export_df["Output"] = outputs_list
    export_df["Remarks/Comments"] = remarks_list

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Triaging Steps")

        worksheet = writer.sheets["Triaging Steps"]

        worksheet.column_dimensions["A"].width = 8
        worksheet.column_dimensions["B"].width = 30
        worksheet.column_dimensions["C"].width = 50
        worksheet.column_dimensions["D"].width = 60
        worksheet.column_dimensions["E"].width = 50
        worksheet.column_dimensions["F"].width = 40

        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    output.seek(0)
    return output.getvalue()


def show_page(session_state, TemplateParser, EnhancedTemplateGenerator):
    """Main page function - UPDATED TO USE INTELLIGENT GENERATOR"""

    selected_alert = session_state.get("triaging_selected_alert", None)

    if selected_alert is None:
        rule_number = session_state.get("triaging_rule_number", None)
        if rule_number:
            selected_alert = {
                "rule": rule_number,
                "rule_number": rule_number,
                "incident": f"TEMPLATE_GEN_{rule_number}",
                "description": f"Template Generation for Rule {rule_number}",
            }
            session_state["triaging_selected_alert"] = selected_alert
        else:
            st.error("❌ No alert selected. Please go back and select an incident.")
            return

    rule_number = selected_alert.get(
        "rule_number", selected_alert.get("rule", "Unknown")
    )

    # Extract rule number for file matching
    rule_num_match = re.search(r"#?(\d+)", rule_number)
    rule_num = (
        rule_num_match.group(1)
        if rule_num_match
        else rule_number.replace("#", "").strip()
    )

    # Find template
    template_dir = "data/triaging_templates"

    if not os.path.exists(template_dir):
        st.warning(f"⚠️ Template directory not found: {template_dir}")
        if st.button("📁 Create Template Directory", type="primary"):
            os.makedirs(template_dir, exist_ok=True)
            st.success("✅ Directory created! Please upload templates.")
            st.stop()
        st.stop()

    all_files = os.listdir(template_dir)
    template_files = [
        f
        for f in all_files
        if rule_num in f and (f.endswith(".csv") or f.endswith(".xlsx"))
    ]

    if not template_files:
        st.error(f"❌ No template found for {rule_number}")
        st.info(f"🔍 Looking for files containing '{rule_num}' in: {template_dir}")
        with st.expander("Available Templates", expanded=True):
            if all_files:
                st.write("Found templates:")
                for f in all_files:
                    st.write(f"- {f}")
            else:
                st.write("No templates found in directory")
        st.stop()

    template_path = os.path.join(template_dir, template_files[0])

    # ✅ CREATE CACHE KEY
    cache_key = _get_enhancement_cache_key(rule_number, template_path)

    # ✅ CHECK CACHE FIRST
    if cache_key in st.session_state:
        cached_data = st.session_state[cache_key]
        session_state.original_steps = cached_data["original_steps"]
        session_state.enhanced_steps = cached_data["enhanced_steps"]
        session_state.excel_template_data = cached_data["excel_template_data"]
        session_state.template_dataframe = cached_data["template_dataframe"]

        _display_enhancement_results(
            session_state,
            rule_number,
            EnhancedTemplateGenerator,
        )
        return

    # ✅ RUN INTELLIGENT ENHANCEMENT (ONLY ONCE)
    st.info(f"📄 Processing template: {template_files[0]}")

    progress_bar = st.progress(0, text="📄 Starting intelligent enhancement...")

    try:
        # STEP 1: Parse Old Template
        progress_bar.progress(10, text="📋 Parsing original template...")

        parser = TemplateParser()

        if template_path.endswith(".csv"):
            original_steps = parser.parse_csv_template(template_path)
        else:
            original_steps = parser.parse_excel_template(template_path)

        if not original_steps:
            st.error("❌ Template parsing failed - no steps extracted")
            st.stop()

        progress_bar.progress(20, text=f"✅ Extracted {len(original_steps)} steps")

        # STEP 2: Use Intelligent Template Generator (NEW!)
        progress_bar.progress(30, text="🧠 Analyzing with AI intelligence...")

        # Initialize the intelligent generator
        from routes.src.template_generator import ImprovedTemplateGenerator

        intelligent_gen = ImprovedTemplateGenerator()

        progress_bar.progress(40, text="📡 Gathering threat intelligence...")

        start_time = time.time()

        # Generate intelligent template
        template_df = intelligent_gen.generate_intelligent_template(
            rule_number=rule_number, original_steps=original_steps
        )

        elapsed = time.time() - start_time

        progress_bar.progress(80, text="📊 Finalizing template...")

        # Convert DataFrame back to dictionary format for display
        enhanced_steps_with_kql = []

        for idx, row in template_df.iterrows():
            # Skip header row
            if pd.isna(row["Step"]) or str(row["Step"]).strip() == "":
                continue

            kql_raw = row["KQL Query"]
            kql_str = str(kql_raw) if pd.notna(kql_raw) else ""
            kql_cleaned = kql_str.strip().lower()

            if pd.notna(kql_raw) and kql_cleaned not in ["nan", "none", ""]:
                final_kql = str(kql_raw)
            else:
                final_kql = ""

            step_dict = {
                "step_name": str(row["Name"]) if pd.notna(row["Name"]) else "",
                "explanation": (
                    str(row["Explanation"]) if pd.notna(row["Explanation"]) else ""
                ),
                "kql_query": final_kql,
                "kql_explanation": (
                    str(row["KQL Explanation"])
                    if pd.notna(row["KQL Explanation"])
                    and str(row["KQL Explanation"]).strip().lower()
                    not in ["nan", "none", ""]
                    else ""
                ),
                "input_required": "",
            }

            enhanced_steps_with_kql.append(step_dict)

        progress_bar.progress(90, text="💾 Caching results...")

        # Export to Excel
        excel_file = intelligent_gen.export_to_excel(template_df, rule_number)

        # ✅ CACHE RESULTS
        st.session_state[cache_key] = {
            "original_steps": original_steps,
            "enhanced_steps": enhanced_steps_with_kql,
            "excel_template_data": excel_file,
            "template_dataframe": template_df,
            "elapsed_time": elapsed,
        }

        session_state.original_steps = original_steps
        session_state.enhanced_steps = enhanced_steps_with_kql
        session_state.excel_template_data = excel_file
        session_state.template_dataframe = template_df

        progress_bar.progress(100, text="✅ Enhancement complete!")

        time.sleep(0.5)
        progress_bar.empty()

        st.success(f"🎉 Intelligent template generated in {elapsed:.1f}s!")

        # Display results
        _display_enhancement_results(
            session_state,
            rule_number,
            EnhancedTemplateGenerator,
        )

    except Exception as e:
        progress_bar.empty()
        st.error(f"❌ Error: {str(e)}")
        with st.expander("🔍 View Error Details"):
            st.code(traceback.format_exc())


def _display_enhancement_results(
    session_state,
    rule_number,
    EnhancedTemplateGenerator,
):
    """Display enhancement results with accordion navigation and completion tracking"""

    enhanced_steps = session_state.enhanced_steps

    st.success(f"✅ Successfully generated {len(enhanced_steps)} enhanced steps")
    st.markdown("---")

    # Initialize state
    if "step_remarks" not in st.session_state:
        st.session_state.step_remarks = {}
    if "step_outputs" not in st.session_state:
        st.session_state.step_outputs = {}
    if "completed_steps" not in st.session_state:
        st.session_state.completed_steps = set()
    if "current_open_step" not in st.session_state:
        st.session_state.current_open_step = 1

    # TWO MAIN TABS
    tab1, tab2 = st.tabs(["📋 Triaging Steps", "📊 Excel Template"])

    # TAB 1: Accordion Steps
    with tab1:
        # Show upload status if exists
        if st.session_state.get("predictions_upload_error"):
            st.error(f"❌ Upload failed: {st.session_state.predictions_upload_error}")
            if st.button("🔄 Retry Upload"):
                del st.session_state.predictions_upload_error
                if st.session_state.get("predictions_excel_data"):
                    _upload_to_predictions_api(
                        st.session_state.predictions_excel_data,
                        st.session_state.predictions_excel_filename,
                    )
                st.rerun()

        if st.session_state.get("show_predictions_unlock_message", False):
            if st.session_state.get("predictions_uploaded"):
                st.success(
                    "✅ Predictions tab unlocked! Switch to the **🔮 Predictions & MITRE** tab to continue."
                )
            else:
                st.warning(
                    "⚠️ Template downloaded but upload to predictions API failed. Check the error above."
                )
            del st.session_state.show_predictions_unlock_message

        # Display all steps as accordions
        for idx, step in enumerate(enhanced_steps):
            step_num = idx + 1
            step_name = step.get("step_name", f"Step {step_num}")

            is_completed = step_num in st.session_state.completed_steps
            is_expanded = (
                step_num == st.session_state.current_open_step and not is_completed
            )
            is_locked = (
                step_num > 1 and (step_num - 1) not in st.session_state.completed_steps
            )

            if is_completed:
                status_icon = "✅"
            elif is_locked:
                status_icon = "🔒"
            else:
                status_icon = "⏳"

            with st.expander(
                f"{status_icon} Step {step_num}: {step_name}", expanded=is_expanded
            ):
                if is_locked:
                    st.warning("🔒 Complete the previous step to unlock this one")
                else:
                    explanation = step.get("explanation", "No explanation provided")
                    st.write(explanation)

                    kql_query = step.get("kql_query", "")
                    kql_explanation = step.get("kql_explanation", "")

                    if kql_query:
                        kql_query = str(kql_query).strip()
                        if kql_query.lower() in ["nan", "none", "n/a", ""]:
                            kql_query = ""

                    # Check for VirusTotal step
                    step_name_lower = step_name.lower()
                    explanation_lower = explanation.lower()

                    is_ip_reputation_step = (
                        ("virustotal" in step_name_lower)
                        or ("virustotal" in explanation_lower)
                        or ("virus total" in step_name_lower)
                        or ("virus total" in explanation_lower)
                        or (
                            contains_ip_not_vip(step_name_lower)
                            and "reputation" in step_name_lower
                        )
                    )

                    # KQL Query Section
                    if kql_query and len(kql_query) > 5:
                        st.markdown("##### 🔎 KQL Query")
                        if kql_explanation and str(kql_explanation).strip() not in [
                            "nan",
                            "none",
                            "n/a",
                            "",
                        ]:
                            st.write(kql_explanation)
                        st.code(kql_query, language="kql")

                    # Output Section
                    if is_ip_reputation_step or (kql_query and len(kql_query) > 5):
                        st.markdown("##### 📊 Output")
                        output_key = f"output_step_{step_num}_{rule_number}"
                        existing_output = st.session_state.step_outputs.get(
                            output_key, ""
                        )

                        # VirusTotal Integration
                        if is_ip_reputation_step:
                            st.info("🎯 **IP Reputation Check using Virus Total & Abuse DB**")

                            default_ip = ""
                            for prev_step in range(1, step_num):
                                prev_output_key = (
                                    f"output_step_{prev_step}_{rule_number}"
                                )
                                prev_output = st.session_state.step_outputs.get(
                                    prev_output_key, ""
                                )

                                import re

                                ip_pattern = r"\b(?:\d{1,3}\.){3}\d{1,3}\b"
                                ips_found = re.findall(ip_pattern, prev_output)
                                if ips_found:
                                    default_ip = ips_found[0]
                                    break

                            col1, col2 = st.columns([4, 1])

                            with col1:
                                ip_input = st.text_input(
                                    "Enter IP Address:",
                                    value=default_ip,
                                    placeholder="e.g., 192.168.1.1",
                                    key=f"vt_ip_step_{step_num}",
                                )

                            with col2:
                                st.write("")
                                st.write("")
                                check_button = st.button(
                                    "🔍 Check",
                                    key=f"vt_check_step_{step_num}",
                                    type="primary",
                                )

                            if check_button and ip_input:
                                with st.spinner("🔍 Checking VirusTotal..."):
                                    vt_result = _check_virustotal_auto(ip_input)

                                if vt_result.get("success"):
                                    formatted_output_ui = vt_result.get(
                                        "formatted_output", ""
                                    )
                                    formatted_output_excel = vt_result.get(
                                        "formatted_output_excel", formatted_output_ui
                                    )

                                    st.session_state.step_outputs[output_key] = (
                                        formatted_output_excel
                                    )

                                    st.markdown(formatted_output_ui)
                                    st.success(
                                        "✅ VirusTotal check complete! Output saved for Excel export."
                                    )

                                    risk_level = vt_result.get("risk_level", "UNKNOWN")
                                    if risk_level == "HIGH":
                                        st.error(
                                            "🚨 **HIGH RISK IP** - Immediate action recommended"
                                        )
                                    elif risk_level == "MEDIUM":
                                        st.warning(
                                            "⚠️ **SUSPICIOUS IP** - Further investigation needed"
                                        )
                                    elif risk_level == "LOW":
                                        st.success(
                                            "✅ **CLEAN IP** - No threats detected"
                                        )
                                else:
                                    st.error(
                                        f"❌ {vt_result.get('error', 'Check failed')}"
                                    )
                                    if vt_result.get("manual_check"):
                                        st.markdown(
                                            vt_result.get("formatted_output", "")
                                        )

                            if existing_output:
                                with st.expander(
                                    "📋 View Saved Output (Excel Format)",
                                    expanded=False,
                                ):
                                    st.text(existing_output)

                        else:
                            # Regular manual output
                            manual_output = st.text_area(
                                "Enter the KQL query output:",
                                value=existing_output,
                                height=200,
                                key=f"output_input_{step_num}",
                                placeholder="Paste the query results here...",
                                on_change=lambda sn=step_num: _save_step_data(
                                    sn, rule_number, "output"
                                ),
                            )
                            st.session_state.step_outputs[output_key] = manual_output

                            if manual_output:
                                st.success(
                                    f"✅ Output saved ({len(manual_output)} characters)"
                                )

                    # Remarks Section
                    st.markdown("##### 💬 Remarks/Comments")
                    remark_key = f"remark_step_{step_num}_{rule_number}"
                    existing_remark = st.session_state.step_remarks.get(remark_key, "")

                    remark = st.text_area(
                        "Add remarks:",
                        value=existing_remark,
                        height=120,
                        key=f"remark_input_{step_num}",
                        placeholder="Enter observations or comments...",
                        label_visibility="collapsed",
                        on_change=lambda sn=step_num: _save_step_data(
                            sn, rule_number, "remark"
                        ),
                    )
                    st.session_state.step_remarks[remark_key] = remark

                    st.markdown("---")

                    # Mark as Complete Button
                    if not is_completed:
                        if st.button(
                            "✅ Mark as Complete",
                            key=f"complete_step_{step_num}",
                            type="primary",
                            width="stretch",
                        ):
                            st.session_state.completed_steps.add(step_num)
                            st.session_state.current_open_step = step_num + 1
                            st.rerun()

        # Final Download Section
        if len(st.session_state.completed_steps) == len(enhanced_steps):
            st.markdown("---")
            st.success("🎉 All steps completed!")

            excel_key = f"final_excel_{rule_number}"
            if excel_key not in st.session_state:
                template_df = session_state.template_dataframe
                excel_with_data = _export_template_with_remarks_and_outputs(
                    template_df,
                    st.session_state.step_remarks,
                    st.session_state.step_outputs,
                    rule_number,
                )
                filename = (
                    f"triaging_template_{rule_number.replace('#', '_')}_complete.xlsx"
                )
                st.session_state[excel_key] = excel_with_data
                st.session_state.final_excel_filename = filename
            else:
                excel_with_data = st.session_state[excel_key]
                filename = st.session_state.final_excel_filename

            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.download_button(
                    label="📥 Download & Proceed to Predictions",
                    data=excel_with_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    width="stretch",
                    key="download_and_proceed",
                    on_click=lambda: _unlock_predictions(
                        excel_with_data, filename, rule_number
                    ),
                )

            if st.session_state.get("show_predictions_unlock_message"):
                if st.session_state.get("predictions_uploaded"):
                    st.success(
                        "✅ Template uploaded to predictions API! Switch to Predictions tab."
                    )
                else:
                    st.error(
                        f"❌ Upload failed: {st.session_state.get('predictions_upload_error', 'Unknown error')}"
                    )
                    if st.button("🔄 Retry Upload"):
                        success = _upload_to_predictions_api(excel_with_data, filename)
                        if success:
                            st.session_state.predictions_uploaded = True
                            st.rerun()

    # TAB 2: Excel Template Preview & Download
    with tab2:
        st.markdown("### 📊 Excel Template Preview")

        if (
            hasattr(session_state, "template_dataframe")
            and session_state.template_dataframe is not None
        ):
            template_df = session_state.template_dataframe
        else:
            st.warning("Template not generated yet")
            return

        st.dataframe(template_df, width="stretch", height=500)

        st.markdown("---")

        col1, col2 = st.columns(2)

        with col1:
            st.download_button(
                label="📥 Download Base Template",
                data=session_state.excel_template_data,
                file_name=f"triaging_template_{rule_number.replace('#', '_')}_base.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
                key="download_base_excel",
            )

        with col2:
            complete_excel_key = f"complete_excel_{rule_number}"
            if complete_excel_key not in st.session_state:
                remarks_dict = st.session_state.step_remarks
                outputs_dict = st.session_state.step_outputs
                st.session_state[complete_excel_key] = (
                    _export_template_with_remarks_and_outputs(
                        template_df, remarks_dict, outputs_dict, rule_number
                    )
                )

            st.download_button(
                label="📥 Download Complete Template",
                data=st.session_state[complete_excel_key],
                file_name=f"triaging_template_{rule_number.replace('#', '_')}_complete.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
                type="primary",
                key="download_complete_excel",
            )
