import pandas as pd
import re
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class EnhancedTemplateGenerator:
    """
    Generates CLEAN Excel templates with NO HARDCODED DATA.
    All KQL queries use placeholders like <USER_EMAIL>, <IP_ADDRESS>, <TIMESPAN>.
    """

    def __init__(self):
        self.template_columns = [
            "Step",
            "Name",
            "Explanation",
            "Input",
            "KQL Query",
            "Execute",
            "Output",
            "Remarks/Comments",
        ]

    def generate_clean_template(
        self, rule_number: str, enhanced_steps: list
    ) -> pd.DataFrame:
        """
        Generate clean Excel template from web-enhanced steps.

        Args:
            rule_number: Rule identifier (e.g., "Rule#183")
            enhanced_steps: List of steps enhanced by web research + LLM

        Returns:
            DataFrame ready for Excel export
        """
        template_rows = []

        # Add rule header row
        header_row = {
            "Step": "",
            "Name": rule_number,
            "Explanation": "Investigation template generated from historical analysis and web research",
            "Input": "",
            "KQL Query": "",
            "Execute": "",
            "Output": "",
            "Remarks/Comments": "",
        }
        template_rows.append(header_row)

        # Add each investigation step
        for i, step in enumerate(enhanced_steps, 1):
            step_row = self._create_clean_step_row(i, step)
            template_rows.append(step_row)

        return pd.DataFrame(template_rows)

    def _create_clean_step_row(self, step_num: int, step: dict) -> dict:
        """
        Create a CLEAN step row with NO HARDCODED VALUES.

        Step structure from LLM/Web enhancement:
        {
            "step_name": "Check User Sign-In Logs",
            "explanation": "Query Azure AD sign-in logs...",
            "input_required": "User principal name (email), Time range",
            "kql_query": "SigninLogs | where UserPrincipalName == '<USER_EMAIL>'...",
            "expected_output": "Clean IP, Known device, MFA satisfied"
        }
        """
        step_name = step.get("step_name", f"Investigation Step {step_num}")
        explanation = step.get("explanation", "")
        input_required = step.get("input_required", "")
        kql_query = step.get("kql_query", "")

        # Clean step name (remove markdown, numbers, emojis)
        clean_name = self._clean_step_name(step_name)

        # Clean explanation (remove markdown, keep concise)
        clean_explanation = self._clean_text(explanation)

        # Clean input requirements
        clean_input = self._clean_text(input_required)

        # Clean KQL query (ensure NO hardcoded values)
        clean_kql = self._ensure_parameterized_kql(kql_query)

        return {
            "Step": step_num,
            "Name": clean_name,
            "Explanation": clean_explanation,
            "Input": clean_input,
            "KQL Query": clean_kql,
            "Execute": "",  # Empty for manual filling
            "Output": "",  # Empty for manual filling
            "Remarks/Comments": "",  # Empty for manual notes
        }

    def _clean_step_name(self, step_name: str) -> str:
        """Clean step name - remove ALL formatting, keep action-focused text"""
        if not step_name:
            return "Investigation Step"

        # Remove markdown
        clean = re.sub(r"\*+", "", step_name)
        clean = re.sub(r"#+", "", clean)
        clean = re.sub(r"[📋✅📊📂🔍🔎⚠️🚨]", "", clean)  # Remove emojis

        # Remove numbering
        clean = re.sub(r"^Step\s*\d+:?\s*", "", clean, flags=re.IGNORECASE)
        clean = re.sub(r"^\d+\.\s*", "", clean)

        # Remove verbose prefixes
        clean = re.sub(
            r"^(Please\s+)?(Perform\s+)?(Complete\s+)?", "", clean, flags=re.IGNORECASE
        )

        # Clean whitespace
        clean = " ".join(clean.strip().split())

        return clean if clean else "Investigation Step"

    def _clean_text(self, text: str) -> str:
        """Remove ALL markdown and formatting, keep plain text"""
        if not text:
            return ""

        # Remove markdown
        clean_text = re.sub(r"\*\*\*+", "", text)
        clean_text = re.sub(r"\*\*", "", clean_text)
        clean_text = re.sub(r"\*", "", clean_text)
        clean_text = re.sub(r"#+\s*", "", clean_text)
        clean_text = re.sub(r"```[a-z]*\n", "", clean_text)
        clean_text = re.sub(r"\n```", "", clean_text)
        clean_text = re.sub(r"`([^`]+)`", r"\1", clean_text)

        # Remove common verbose prefixes
        clean_text = re.sub(
            r"^(Explanation:|Description:|Instructions:)\s*",
            "",
            clean_text,
            flags=re.IGNORECASE,
        )

        # Clean whitespace
        clean_text = re.sub(r"\s+", " ", clean_text)
        clean_text = clean_text.strip()

        return clean_text

    def _ensure_parameterized_kql(self, kql: str) -> str:
        """
        CRITICAL: Ensure KQL query has NO HARDCODED VALUES.
        Replace any hardcoded data with placeholders.
        """
        if not kql or kql.strip() == "" or kql.strip().upper() == "N/A":
            return ""

        # Remove markdown code blocks
        clean_kql = re.sub(r"```[a-z]*\s*\n?", "", kql)
        clean_kql = re.sub(r"\n?```", "", clean_kql)

        # Replace any hardcoded emails with placeholder
        clean_kql = re.sub(
            r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", "<USER_EMAIL>", clean_kql
        )

        # Replace any hardcoded IPs with placeholder
        clean_kql = re.sub(
            r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", "<IP_ADDRESS>", clean_kql
        )

        # Replace any hardcoded device names with placeholder
        clean_kql = re.sub(
            r'(DeviceId|DeviceName)\s*==\s*"[^"]+"', r'\1 == "<DEVICE_ID>"', clean_kql
        )

        # Ensure time ranges use placeholders
        if "ago(" in clean_kql and "d)" in clean_kql:
            # Already parameterized, keep as is
            pass
        elif "TimeGenerated" in clean_kql and "ago" not in clean_kql:
            # Add ago() if missing
            clean_kql = re.sub(
                r'(TimeGenerated\s*>\s*)["\'][\d\-:TZ]+["\']',
                r"\1ago(<TIMESPAN>)",
                clean_kql,
            )

        # Clean up whitespace but preserve line structure
        lines = [line.strip() for line in clean_kql.split("\n") if line.strip()]
        clean_kql = "\n".join(lines)

        return clean_kql.strip()

    def export_to_excel(self, df: pd.DataFrame, rule_number: str) -> BytesIO:
        """Export DataFrame to professionally formatted Excel"""
        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Triaging_Template", index=False)

            # Format the worksheet
            workbook = writer.book
            worksheet = writer.sheets["Triaging_Template"]
            self._format_excel_worksheet(worksheet, df)

        output.seek(0)
        return output

    def _format_excel_worksheet(self, worksheet, df):
        """Apply professional formatting to Excel worksheet"""

        # Header formatting
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # Apply header formatting
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Column widths
        column_widths = {
            "A": 8,  # Step
            "B": 35,  # Name
            "C": 50,  # Explanation
            "D": 30,  # Input
            "E": 60,  # KQL Query (wider for full queries)
            "F": 12,  # Execute
            "G": 12,  # Output
            "H": 12,  # Remarks/Comments
        }

        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        # Cell formatting
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        cell_alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")

        # Apply to all data cells
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = cell_alignment

                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color="F9F9F9", end_color="F9F9F9", fill_type="solid"
                    )

        # Highlight header row (rule description)
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=2, column=col_idx)
            cell.fill = PatternFill(
                start_color="E8F4F8", end_color="E8F4F8", fill_type="solid"
            )
            cell.font = Font(bold=True, size=11)

        # Freeze header row
        worksheet.freeze_panes = "A2"
