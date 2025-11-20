import re
import os
import time
import pandas as pd
from io import BytesIO
from typing import List, Dict
from dotenv import load_dotenv
from crewai import LLM, Agent, Task, Crew
from routes.src.utils import _strip_step_number_prefix
from api_client.analyzer_api_client import get_analyzer_client
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from difflib import SequenceMatcher
from routes.src.api_kql_generation import EnhancedKQLGenerator

load_dotenv()

# Disable verbose LiteLLM logging
os.environ["LITELLM_LOG"] = "ERROR"
import logging
logging.getLogger("LiteLLM").setLevel(logging.ERROR)
logging.getLogger("httpx").setLevel(logging.ERROR)


class ImprovedTemplateGenerator:
    def __init__(self):
        # Initialize LLM for non-KQL tasks with multi-key support
        api_keys = [
            os.getenv("GOOGLE_API_KEY_1", os.getenv("GOOGLE_API_KEY")),
            os.getenv("GOOGLE_API_KEY_2"),
            os.getenv("GOOGLE_API_KEY_3"),
            os.getenv("GOOGLE_API_KEY_4"),
            os.getenv("GOOGLE_API_KEY_5"),
            os.getenv("GOOGLE_API_KEY_6"),
            os.getenv("GOOGLE_API_KEY_7")
        ]
        api_keys = [key for key in api_keys if key]  # Filter out None values
        
        if api_keys:
            # Use first available API key
            gemini_key = api_keys[0]
            self.llm = LLM(
                model="gemini/gemini-1.5-flash", api_key=gemini_key, temperature=0.3
            )
            print(f"✅ Using Gemini for template generation (key {1}/{len(api_keys)})")
        else:
            ollama_model = os.getenv("OLLAMA_CHAT", "ollama/qwen2.5:3b")
            if not ollama_model.startswith("ollama/"):
                ollama_model = f"ollama/{ollama_model}"
            self.llm = LLM(
                model=ollama_model, base_url="http://localhost:11434", temperature=0.3
            )
            print(f"✅ Using {ollama_model} for template generation")

        # Initialize enhanced KQL generator (will be updated with alert source type when needed)
        self.kql_generator = None

        self.template_columns = [
            "Step",
            "Name",
            "Explanation",
            "KQL Query",
            "KQL Explanation",
            "Execute",
            "Output",
            "Remarks/Comments",
        ]

        # Initialize API Client for getting the Technical Overview
        try:
            self.analyzer_client = get_analyzer_client()
        except Exception as e:
            print(f"⚠️ Could not initialize Analyzer API Client: {e}")
            self.analyzer_client = None

    def generate_intelligent_template(
        self, rule_number: str, original_steps: List[Dict], rule_context: str = "", alert_data: dict = None
    ) -> pd.DataFrame:
        """
        Generate intelligent template with support for manual alert analysis

        ✅ UPDATED: Detects manual alerts and routes to generate_from_manual_analysis()
        """
        print(f"\n{'='*80}")
        print(f"🧠 INTELLIGENT TEMPLATE GENERATION")
        print(f"{'='*80}\n")
        
        # Store alert_data for profile building
        self._current_alert_data = alert_data

        # ✅ CHECK IF THIS IS A MANUAL ALERT GENERATION
        if rule_number == "MANUAL_GEN" and not original_steps:
            print(f"🔍 Manual alert detected - using direct analysis mode")
            # rule_context contains the analysis text
            alert_name = (
                rule_context.split("\n")[0][:100] if rule_context else "Manual Alert"
            )
            return self.generate_from_manual_analysis(
                alert_name, rule_context, rule_number
            )

        # ✅ ORIGINAL FLOW: Template-based generation
        print(f"Processing template-based generation for {rule_number}")

        start_time = time.time()

        # Header row
        template_rows = []
        header_row = {col: "" for col in self.template_columns}
        header_row["Name"] = rule_number
        template_rows.append(header_row)

        # STEP 1: BUILD INVESTIGATION PROFILE
        print(f"\n📊 PHASE 1: Building investigation profile...")
        from routes.src.new_steps.investigation_profile import (
            InvestigationProfileBuilder,
        )

        profile_builder = InvestigationProfileBuilder()
        # Pass alert_data to profile builder to extract alert_source_type
        alert_data_for_profile = None
        if hasattr(self, '_current_alert_data'):
            alert_data_for_profile = self._current_alert_data
        profile = profile_builder.build_profile(rule_number, rule_context, alert_data_for_profile)

        profile["existing_step_names"] = [
            s.get("step_name", "") for s in original_steps
        ]

        print(f"   ✅ Profile complete:")
        print(f"      - MITRE Techniques: {len(profile['mitre_techniques'])}")
        print(f"      - Threat Actors: {len(profile['threat_actors'])}")
        print(f"      - Investigation Focus: {profile['investigation_focus']}")

        # STEP 2: GENERATE INVESTIGATION STEPS
        print(f"\n🤖 PHASE 2: Generating investigation steps...")
        from routes.src.new_steps.step_library import InvestigationStepLibrary

        step_library = InvestigationStepLibrary()
        
        # Add retry logic for step generation
        max_retries = 3
        generated_steps = []
        
        for attempt in range(max_retries):
            try:
                generated_steps = step_library.generate_investigation_steps(profile)
                print(f"   ✅ Generated {len(generated_steps)} investigation steps")
                break
            except Exception as e:
                print(f"   ⚠️ Step generation attempt {attempt + 1} failed: {str(e)[:100]}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
                else:
                    print(f"   ❌ Using fallback steps due to LLM failure")
                    # Create basic fallback steps
                    generated_steps = [
                        {"step_name": "Manual Investigation", "explanation": "Conduct manual investigation due to service unavailability", "priority": "HIGH"},
                        {"step_name": "Document Findings", "explanation": "Document all investigation findings", "priority": "MEDIUM"}
                    ]

        # STEP 3: FILTER & MERGE STEPS
        print(f"\n🔄 PHASE 3: Merging with original template...")

        investigative_original = self._filter_investigative_steps(original_steps)
        print(f"   ✅ Original steps (investigative): {len(investigative_original)}")

        from routes.src.new_steps.step_merger import InvestigationStepMerger

        merger = InvestigationStepMerger()
        merged_steps, merge_report = merger.merge_steps(
            investigative_original, generated_steps, profile
        )

        merger.print_merge_report(merge_report)

        # STEP 4: ADD KQL & CONVERT TO TEMPLATE ROWS
        print(f"\n⚙️ PHASE 4: Generating KQL queries and finalizing...")
        
        template_rows.extend(
            self._process_merged_steps_with_kql(merged_steps, rule_number, profile)
        )

        template_rows = self._deduplicate_kql_queries(template_rows)

        elapsed = time.time() - start_time
        print(f"\n{'='*80}")
        print(
            f"✅ COMPLETED in {elapsed:.1f}s: {len(template_rows)-1} investigation steps"
        )
        print(f"{'='*80}\n")

        return pd.DataFrame(template_rows)

    def _process_merged_steps_with_kql(
        self, merged_steps: List[Dict], rule_number: str, profile: Dict
    ) -> List[Dict]:
        """
        Process merged steps and generate KQL for each

        UPDATED: Extract datetime from profile and pass to KQL generator
        """
        template_rows = []

        # NEW: Extract reference datetime from profile or alert_data
        reference_datetime_obj = None
        if profile.get("reference_datetime_obj"):
            reference_datetime_obj = profile["reference_datetime_obj"]
            
        # Initialize KQL generator with alert source type if available
        alert_source_type = profile.get("alert_source_type", "")
        print(f"   🔍 DEBUG: profile alert_source_type: '{alert_source_type}'")
        if not self.kql_generator:
            print(f"   🔍 DEBUG: Creating new KQL generator with alert_source_type: '{alert_source_type}'")
            self.kql_generator = EnhancedKQLGenerator(alert_source_type=alert_source_type)
        else:
            print(f"   🔍 DEBUG: Updating existing KQL generator with alert_source_type: '{alert_source_type}'")
            # Update existing generator with alert source type
            self.kql_generator.alert_source_type = alert_source_type

        # Separate original and AI-generated steps
        original_steps = []
        ai_generated_steps = []

        for step in merged_steps:
            if step.get("source") == "original_template":
                original_steps.append(step)
            else:
                ai_generated_steps.append(step)

        print(f"   Original steps: {len(original_steps)}")
        print(f"   AI-generated steps: {len(ai_generated_steps)}")

        # FIRST: Process ORIGINAL steps - keep ALL regardless of KQL
        print(f"\n   OK: Processing ORIGINAL steps (keeping all)...")
        original_processed = 0
        for idx, step in enumerate(original_steps, 1):
            step_name = _strip_step_number_prefix(step.get("step_name", ""))
            explanation = step.get("explanation", "")
            source = step.get("source", "unknown")
            priority = step.get("priority", "MEDIUM")
            confidence = step.get("confidence", "MEDIUM")

            print(f"\n      Step {idx}: {step_name}")
            print(f"         Source: {source} | Priority: {priority}")

            # Try to generate KQL for original steps too
            kql_query = ""
            kql_explanation = ""

            # Explicit check for external tool usage
            tool_used = step.get("tool", "").lower()

            # Check if this is a VIP verification step
            is_vip_step = any(
                kw in step_name.lower() or kw in explanation.lower()
                for kw in [
                    "vip",
                    "executive",
                    "high-priority",
                    "privileged account",
                    "account status",
                ]
            )

            # UPDATED: Check tool first, then general _needs_kql
            if tool_used in ["virustotal", "abuseipdb"]:
                print(f"         INFO: Skipping KQL for External Tool: {tool_used}")
            elif is_vip_step:
                print(
                    f"         NOTE: VIP Step Detected - Generating template query..."
                )

                # PASS reference_datetime_obj to KQL generator
                kql_query, kql_explanation = self.kql_generator.generate_kql_query(
                    step_name="VIP User Verification",
                    explanation="Verify if affected user is VIP or executive account",
                    rule_context=profile.get("technical_overview", ""),
                    reference_datetime_obj=reference_datetime_obj,
                )

                if kql_query and len(kql_query.strip()) > 30:
                    print(
                        f"         OK: VIP KQL template generated ({len(kql_query)} chars)"
                    )
                else:
                    print(
                        f"         WARNING: VIP KQL generation failed - using fallback"
                    )
                    from routes.src.hardcode_kql_queries import HardcodedKQLQueries

                    kql_query = HardcodedKQLQueries.VIP_ACCOUNT_VERIFICATION
                    kql_explanation = "VIP/Executive account verification query"

            elif self._needs_kql(step_name, explanation):
                print(f"         INFO: Generating KQL...")

                # PASS reference_datetime_obj to KQL generator
                kql_query, kql_explanation = self.kql_generator.generate_kql_query(
                    step_name=step_name,
                    explanation=explanation,
                    rule_context=profile.get("technical_overview", ""),
                    reference_datetime_obj=reference_datetime_obj,
                )

                if kql_query and len(kql_query.strip()) > 30:
                    print(f"         OK: KQL generated ({len(kql_query)} chars)")
                else:
                    print(
                        f"         INFO: No KQL generated (External Tool/Manual/Fallback)"
                    )
                    kql_query = ""
            else:
                print(f"         INFO: Step doesn't require KQL (manual/closure)")

            # Build template row - KEEP ORIGINAL EVEN IF NO KQL
            row = {
                "Step": str(idx),
                "Name": step_name,
                "Explanation": self._build_enhanced_explanation(
                    explanation,
                    step.get("relevance", ""),
                    profile.get("alert_name", ""),
                ),
                "KQL Query": kql_query,
                "KQL Explanation": kql_explanation,
                "Execute": "",
                "Output": "",
                "Remarks/Comments": f"[{source.upper()}] {priority}",
            }

            template_rows.append(row)
            original_processed += 1

        print(f"\n   OK: Kept all {original_processed} ORIGINAL steps")

        # SECOND: Process AI-GENERATED steps - FILTER if no KQL
        print(f"\n   OK: Processing AI-GENERATED steps (filtering if no KQL)...")
        ai_processed = 0
        ai_skipped = 0

        # Adjusting the step numbering based on processed original steps
        starting_idx = original_processed + 1

        for idx, step in enumerate(ai_generated_steps, starting_idx):
            step_name = _strip_step_number_prefix(step.get("step_name", ""))
            explanation = step.get("explanation", "")
            source = step.get("source", "unknown")
            priority = step.get("priority", "MEDIUM")
            confidence = step.get("confidence", "MEDIUM")

            tool_used = step.get("tool", "").lower()

            # FOR AI-GENERATED: Check tool first (e.g., VirusTotal)
            if tool_used in ["virustotal", "abuseipdb"]:
                kql_query = ""
                kql_explanation = "Requires manual checking using external tools (VirusTotal, AbuseIPDB) or the integrated IP reputation checker in the triaging app."
                print(f"      OK: Keeping AI step (External Tool): {step_name}")

            elif not self._needs_kql(step_name, explanation):
                print(
                    f"      NOTE: Skipping AI step (no KQL needed/manual): {step_name}"
                )
                ai_skipped += 1
                continue

            # If KQL needed and not an external tool, try to generate it
            else:
                print(f"\n      Step {idx}: {step_name}")
                print(f"         Source: {source} | Priority: {priority}")
                print(f"         INFO: Generating KQL...")

                # PASS reference_datetime_obj to KQL generator
                kql_query, kql_explanation = self.kql_generator.generate_kql_query(
                    step_name=step_name,
                    explanation=explanation,
                    rule_context=profile.get("technical_overview", ""),
                    reference_datetime_obj=reference_datetime_obj,
                )

                # FOR AI-GENERATED: Skip if KQL generation failed
                if not (kql_query and len(kql_query.strip()) > 30):
                    print(f"         NOTE: Skipping (KQL generation failed)")
                    ai_skipped += 1
                    continue

                print(f"         OK: KQL generated ({len(kql_query)} chars)")
                print(f"         INFO: Explanation: {kql_explanation[:80]}...")

            # Build template row for AI-generated step
            row = {
                "Step": str(idx),
                "Name": step_name,
                "Explanation": self._build_enhanced_explanation(
                    explanation,
                    step.get("relevance", ""),
                    profile.get("alert_name", ""),
                ),
                "KQL Query": kql_query,
                "KQL Explanation": kql_explanation,
                "Execute": "",
                "Output": "",
                "Remarks/Comments": f"[{source.upper()}] {priority}",
            }

            template_rows.append(row)
            ai_processed += 1

        print(f"\n   OK: Processed {ai_processed} AI-GENERATED steps")
        print(f"   NOTE: Skipped {ai_skipped} AI-GENERATED steps (no KQL)")
        print(f"\n   RESULT: Final template: {len(template_rows)} total steps")
        print(f"      - {original_processed} from ORIGINAL")
        print(f"      - {ai_processed} from AI-GENERATED")

        # NEW: Deduplicate KQL queries before returning
        print(f"\n   INFO: Final KQL deduplication...")
        template_rows = self._deduplicate_kql_in_rows(template_rows)

        return template_rows

    def generate_from_manual_analysis(
        self,
        alert_name: str,
        analysis_text: str,
        rule_number: str = "MANUAL_GEN",
        alert_data: dict = None,
    ) -> pd.DataFrame:
        """
        Generate manual alert investigation template

        UPDATED: Extract reference_datetime_obj and PASS IT through entire pipeline
        """
        print(f"\n{'='*80}")
        print(f"INFO: GENERATING MANUAL ALERT TEMPLATE")
        print(f"Alert: {alert_name}")
        print(f"Has alert_data: {'YES' if alert_data else 'NO'}")
        print(f"{'='*80}\n")

        start_time = time.time()

        # NEW: Extract reference datetime from alert_data
        reference_datetime_obj = None
        if alert_data:
            try:
                full_alert = alert_data.get("full_alert", {})
                if isinstance(full_alert, dict):
                    props = full_alert.get("properties", {})
                    time_str = props.get("timeGenerated")
                    if time_str:
                        from datetime import datetime

                        reference_datetime_obj = datetime.fromisoformat(
                            time_str.replace("Z", "+00:00")
                        )
                        print(
                            f"OK: Reference DateTime extracted: {reference_datetime_obj}"
                        )
            except Exception as e:
                print(f"WARNING: Could not extract reference datetime: {e}")

        # Extract technical overview from analysis_text if available
        technical_overview = ""
        if analysis_text and "## 1. TECHNICAL OVERVIEW" in analysis_text:
            sections = analysis_text.split("##")
            for section in sections:
                if "1. TECHNICAL OVERVIEW" in section:
                    technical_overview = section.replace("1. TECHNICAL OVERVIEW", "").strip()
                    break
        elif alert_data and alert_data.get("technical_overview"):
            technical_overview = alert_data.get("technical_overview")
        
        print(f"Technical overview extracted: {len(technical_overview)} chars" if technical_overview else "No technical overview found")

        # Initialize KQL generator with alert source type for manual analysis
        alert_source_type = alert_data.get("alert_source_type", "") if alert_data else ""
        print(f"🔍 DEBUG: Manual analysis - alert_source_type: '{alert_source_type}'")
        if not self.kql_generator:
            print(f"🔍 DEBUG: Creating KQL generator for manual analysis with alert_source_type: '{alert_source_type}'")
            self.kql_generator = EnhancedKQLGenerator(alert_source_type=alert_source_type)
        else:
            print(f"🔍 DEBUG: Updating KQL generator for manual analysis with alert_source_type: '{alert_source_type}'")
            # Update existing generator with alert source type
            self.kql_generator.alert_source_type = alert_source_type
        
        # Use the updated step library to generate steps from analysis
        from routes.src.new_steps.step_library import InvestigationStepLibrary

        step_library = InvestigationStepLibrary()

        # Generate steps with enhanced context including technical overview
        max_retries = 3
        generated_steps = []
        
        for attempt in range(max_retries):
            try:
                # Enhanced context for step generation
                enhanced_context = {
                    "alert_name": alert_name,
                    "analysis_text": analysis_text,
                    "technical_overview": technical_overview,
                    "rule_number": rule_number,
                    "alert_data": alert_data,
                }
                
                generated_steps = step_library.generate_steps_from_manual_analysis(
                    alert_name=alert_name,
                    analysis_text=analysis_text,
                    rule_number=rule_number,
                    alert_data=alert_data,
                    technical_overview=technical_overview,  # Pass technical overview
                )
                break
            except Exception as e:
                print(f"   ⚠️ Manual analysis attempt {attempt + 1} failed: {str(e)[:100]}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
                else:
                    print(f"   ❌ Using fallback steps due to LLM failure")
                    # Create enhanced fallback steps using technical overview if available
                    if technical_overview:
                        generated_steps = [
                            {"step_name": "Verify Alert Details", "explanation": f"Review and verify the alert details based on: {technical_overview[:200]}...", "priority": "HIGH"},
                            {"step_name": "Check User Activity", "explanation": "Investigate user authentication and activity patterns related to this security event", "priority": "HIGH"},
                            {"step_name": "VIP User Verification", "explanation": "Check if any affected users are VIP or executive accounts requiring special handling", "priority": "HIGH", "input_required": "vip_user_list"},
                            {"step_name": "IP Reputation Check", "explanation": "Verify reputation of source IP addresses using VirusTotal and check for VPN/proxy usage", "priority": "MEDIUM"},
                            {"step_name": "Document Findings", "explanation": "Document all investigation findings and provide recommendations", "priority": "MEDIUM"}
                        ]
                    else:
                        generated_steps = [
                            {"step_name": "Verify Alert Details", "explanation": "Review and verify the alert details and affected systems", "priority": "HIGH"},
                            {"step_name": "Check User Activity", "explanation": "Investigate user authentication and activity patterns", "priority": "HIGH"},
                            {"step_name": "VIP User Verification", "explanation": "Check if any affected users are VIP or executive accounts", "priority": "HIGH", "input_required": "vip_user_list"},
                            {"step_name": "IP Reputation Check", "explanation": "Verify reputation of source IP addresses", "priority": "MEDIUM"},
                            {"step_name": "Document Findings", "explanation": "Document all investigation findings", "priority": "MEDIUM"}
                        ]

        # Convert to template rows
        template_rows = []

        # Add header row
        header_row = {col: "" for col in self.template_columns}
        header_row["Name"] = f"Manual Analysis: {alert_name}"
        template_rows.append(header_row)

        # Add investigation steps WITH KQL GENERATION USING reference_datetime_obj
        print(f"\nINFO: Processing {len(generated_steps)} generated steps...")
        for idx, step in enumerate(generated_steps, 1):
            step_name = _strip_step_number_prefix(step.get("step_name", ""))
            explanation = step.get("explanation", "")
            relevance = step.get("relevance", "")
            tool = step.get("tool", "")
            existing_kql = step.get("kql_query", "")

            # Build enhanced explanation with relevance and technical context
            full_explanation = explanation
            if relevance:
                full_explanation += f"\n\nWHY THIS MATTERS: {relevance}"
            
            # Add technical context if this is a manual alert with technical overview
            if technical_overview and idx <= 2:  # Only for first 2 steps to avoid repetition
                full_explanation += f"\n\nTECHNICAL CONTEXT: {technical_overview[:300]}..."

            # UPDATED: Generate or enhance KQL if not already present
            # Pass reference_datetime_obj to ensure proper time windows
            kql_query = ""
            kql_explanation = ""

            if existing_kql and len(existing_kql.strip()) > 30:
                # Already has KQL - use it but optionally standardize if needed
                print(f"   Step {idx}: {step_name[:60]} - has existing KQL")
                kql_query = existing_kql
                kql_explanation = step.get("kql_explanation", "")
            else:
                # No KQL yet - try to generate using KQL generator with reference_datetime_obj
                print(f"   Step {idx}: {step_name[:60]} - generating KQL...")

                if self._needs_kql(step_name, explanation):
                    try:
                        kql_query, kql_explanation = (
                            self.kql_generator.generate_kql_query(
                                step_name=step_name,
                                explanation=explanation,
                                rule_context=alert_name,
                                reference_datetime_obj=reference_datetime_obj,  # PASS IT HERE
                            )
                        )

                        if kql_query and len(kql_query.strip()) > 30:
                            print(f"      OK: KQL generated ({len(kql_query)} chars)")
                        else:
                            print(f"      INFO: No KQL generated for this step")
                            kql_query = ""

                    except Exception as e:
                        print(f"      WARNING: KQL generation failed: {str(e)[:100]}")
                        kql_query = ""
                else:
                    print(f"      INFO: Step doesn't need KQL (manual/external tool)")

            row = {
                "Step": str(idx),
                "Name": step_name,
                "Explanation": self._enforce_length(full_explanation, max_sentences=5),
                "KQL Query": kql_query,
                "KQL Explanation": kql_explanation,
                "Execute": "",
                "Output": "",
                "Remarks/Comments": f"[MANUAL] {step.get('priority', 'MEDIUM')} | Tool: {tool if tool else 'KQL'} | Enhanced with AI analysis",
            }

            template_rows.append(row)

        # Create initial DataFrame
        template_df = pd.DataFrame(template_rows)

        # UPDATED: INJECT REAL DATA INTO KQL QUERIES IF ALERT_DATA AVAILABLE
        # This also uses reference_datetime_obj from alert_data
        if alert_data:
            print(f"\n{'='*80}")
            print(f"INFO: INJECTING REAL ALERT DATA INTO KQL QUERIES")
            print(f"{'='*80}")

            # Create a modified alert_data with reference_datetime_obj for injection
            enriched_alert_data = alert_data.copy()
            if reference_datetime_obj:
                enriched_alert_data["reference_datetime_obj"] = reference_datetime_obj

            template_df = self.inject_alert_data_into_template(
                template_df, enriched_alert_data
            )
        else:
            print(
                f"WARNING: No alert_data provided - KQL queries will contain placeholders"
            )

        elapsed = time.time() - start_time

        print(f"\n{'='*80}")
        print(
            f"OK: COMPLETED in {elapsed:.1f}s: {len(template_rows)-1} investigation steps"
        )
        print(f"{'='*80}\n")

        return template_df

    def inject_alert_data_into_template(
        self, template_df: pd.DataFrame, alert_data: dict
    ) -> pd.DataFrame:
        """
        Inject real alert data into KQL queries after template generation

        UPDATED: Use reference_datetime_obj from alert_data if available

        Args:
            template_df: Generated template DataFrame
            alert_data: Alert data containing entities, timestamps, etc.

        Returns:
            DataFrame with injected KQL queries
        """
        print(f"\n{'='*80}")
        print(f"INFO: KQL DATA INJECTION PHASE")
        print(f"{'='*80}\n")

        from routes.src.kql_template_injector import TemplateKQLInjector

        try:
            # UPDATED: Pass reference_datetime_obj if available
            injector = TemplateKQLInjector(alert_data)

            print(f"INFO: Extracted Entities:")
            print(f"   Users: {len(injector.users)}")
            for user in injector.users:
                print(f"      - {user}")
            print(f"   IPs: {len(injector.ips)}")
            for ip in injector.ips:
                print(f"      - {ip}")
            print(f"   Hosts: {len(injector.hosts)}")
            for host in injector.hosts:
                print(f"      - {host}")
            print(f"   Reference DateTime: {injector.reference_datetime}")

            # Inject data into template
            injected_df = injector.inject_template_dataframe(template_df)

            print(f"\nOK: KQL Injection Complete!")
            print(f"{'='*80}\n")

            return injected_df

        except Exception as e:
            print(f"WARNING: KQL Injection failed: {str(e)}")
            import traceback

            print(f"Error trace:\n{traceback.format_exc()}")
            # Return original template if injection fails
            return template_df

    def _deduplicate_kql_in_rows(self, template_rows: List[Dict]) -> List[Dict]:
        """
        Remove template rows with duplicate KQL queries
        """
        from difflib import SequenceMatcher

        seen_queries = {}
        deduplicated = []
        removed_count = 0

        for row in template_rows:
            kql = row.get("KQL Query", "").strip()
            step_name = row.get("Name", "")

            # Keep non-KQL rows (headers, external tools)
            if not kql or len(kql) < 30:
                deduplicated.append(row)
                continue

            # Normalize for comparison
            normalized = re.sub(r"\s+", " ", kql.lower())

            # Check similarity
            is_duplicate = False
            for seen_name, seen_query in seen_queries.items():
                similarity = SequenceMatcher(None, normalized, seen_query).ratio()
                if similarity > 0.85:
                    print(f"      ⏭️  Removing duplicate: {step_name[:60]}")
                    print(f"         (same as: {seen_name[:60]})")
                    is_duplicate = True
                    removed_count += 1
                    break

            if not is_duplicate:
                seen_queries[step_name] = normalized
                deduplicated.append(row)

        # ✅ RENUMBER STEPS after deduplication
        step_counter = 1
        for row in deduplicated:
            if row.get("Step") and str(row["Step"]).strip().isdigit():
                row["Step"] = str(step_counter)
                step_counter += 1

        print(f"   ✅ Removed {removed_count} duplicate queries")
        print(f"   ✅ Final template: {len(deduplicated)-1} unique steps")

        return deduplicated

    def _enhance_step_explanation(self, explanation: str) -> str:
        """
        Enhance step explanation using LLM - make it clear and concise
        This is used as a helper when explanation is too vague
        """
        # ... (rest of _enhance_step_explanation remains the same)
        if not explanation or len(explanation.strip()) == 0:
            return ""

        # If explanation is already reasonable, just enforce length
        if len(explanation) > 30 and not self._is_vague(explanation):
            return self._enforce_length(explanation, max_sentences=3)

        # If explanation is too vague or short, use LLM to enhance it
        prompt = f"""Improve this SOC investigation instruction to be clear and actionable in simple language.

    Current instruction: {explanation[:200]}

    Requirements:
    1. Write in 3 parts:
    - WHAT to check: "This step examines..."
    - WHY it matters: "This is important because..."
    - WHAT to look for: "Look for indicators like..."
    2. Use simple, clear language (explain like talking to a junior analyst)
    3. Be specific about what data/logs to examine
    4. Maximum 3-4 sentences total
    5. Use imperative voice (e.g., "Check", "Review", "Verify")

    Output ONLY the improved instruction (no extra text):"""

        try:
            result = self._quick_llm_call(prompt, max_tokens=150)
            enhanced = self._aggressive_clean(result)

            if len(enhanced) > 40:  # Minimum viable explanation
                return self._enforce_length(enhanced, max_sentences=4)
        except Exception as e:
            print(f"   ⚠️ Explanation enhancement failed: {str(e)[:50]}")

        # Fallback to original if LLM fails
        return self._enforce_length(explanation, max_sentences=3)

    def _build_enhanced_explanation(
        self, explanation: str, relevance: str, alert_name: str
    ) -> str:
        """
        Build comprehensive explanation that includes both HOW and WHY
        Uses LLM enhancement if explanation is vague
        """
        if not explanation:
            return ""

        # Step 1: Enhance the main explanation if it's vague or unclear
        main_explanation = explanation.strip()

        if self._is_vague(main_explanation) or len(main_explanation) < 50:
            # Use LLM to improve vague explanations
            main_explanation = self._enhance_step_explanation(main_explanation)
        else:
            # Just enforce length for clear explanations
            main_explanation = self._enforce_length(main_explanation, max_sentences=3)

        # Step 2: Add relevance context if available
        if relevance and len(relevance.strip()) > 20:
            # Clean up relevance text
            relevance_clean = self._aggressive_clean(relevance.strip())
            relevance_clean = self._enforce_length(relevance_clean, max_sentences=2)

            # Combine with visual separator
            enhanced = f"{main_explanation}\n\n💡 WHY THIS MATTERS: {relevance_clean}"
        else:
            enhanced = main_explanation

        # Step 3: Ensure overall length is reasonable
        return self._enforce_length(enhanced, max_sentences=5)

    def _needs_kql(self, step_name: str, explanation: str) -> bool:
        """Determine if step needs KQL query - MORE INCLUSIVE"""
        combined = f"{step_name} {explanation}".lower()

        # ❌ Skip external tools, reporting, and closure steps
        # This list must be synchronized with what is explicitly NOT given KQL/data_source by the LLM
        skip_keywords = [
            "virustotal",
            "virus total",
            "abuseipdb",
            "abuse",
            "document",
            "close incident",
            "escalate",
            "inform",
            "notify",
            "report",
            "classify",  # e.g. "Classify as TP/FP"
            "manual",  # e.g. "Manual investigation"
        ]

        if any(keyword in combined for keyword in skip_keywords):
            return False

        # ✅ Include all data investigation types
        needs_keywords = [
            "sign-in",
            "signin",
            "login",
            "audit",
            "logs",
            "query",
            "check user",
            "verify user",
            "review",
            "analyze",
            "investigate",
            "count",
            "gather",
            "extract",
            "device",
            "endpoint",
            "role",
            "permission",
            "assignment",
            "group",
            "membership",
            "mfa",
            "authentication",
            "location",
            "oauth",
            "grant",
        ]

        return any(keyword in combined for keyword in needs_keywords)

    def _filter_investigative_steps(self, steps: List[Dict]) -> List[Dict]:
        """Filter non-investigative steps (remediation, closure, etc.)"""
        from routes.src.new_steps.step_merger import InvestigationStepMerger

        merger = InvestigationStepMerger()
        return merger._filter_investigative_steps(steps)

    def _deduplicate_kql_queries(self, template_rows: List[Dict]) -> List[Dict]:
        """Remove steps with duplicate or nearly-identical KQL queries"""
        seen_queries = {}
        deduplicated = []

        for row in template_rows:
            kql = row.get("KQL Query", "").strip()

            # Always keep steps without KQL (header, manual steps)
            if not kql or len(kql) < 20:
                deduplicated.append(row)
                continue

            # Normalize KQL for comparison (remove whitespace, lowercase)
            normalized = re.sub(r"\s+", " ", kql.lower())

            # Check if we've seen this query before
            is_duplicate = False
            for seen_query in seen_queries.values():
                # If queries are 90%+ similar, consider duplicate
                similarity = SequenceMatcher(None, normalized, seen_query).ratio()
                if similarity > 0.9:
                    is_duplicate = True
                    break

            if not is_duplicate:
                seen_queries[row.get("Name", "")] = normalized
                deduplicated.append(row)
            else:
                print(f"   🗑️ Removed duplicate KQL step: {row.get('Name', '')}")

        return deduplicated

    def _enforce_length(self, text: str, max_sentences: int = 3) -> str:
        """Enforce maximum sentence length"""
        if not text:
            return text

        sentences = re.split(r"(?<=[.!?])\s+", text.strip())

        if len(sentences) > max_sentences:
            limited = " ".join(sentences[:max_sentences])
            return limited

        return text

    def _quick_llm_call(self, prompt: str, max_tokens: int = 100) -> str:
        """Make quick LLM call with retry logic"""
        max_retries = 3
        
        for attempt in range(max_retries):
            try:
                agent = Agent(
                    role="SOC Analyst",
                    goal="Generate concise security content",
                    backstory="Expert security analyst",
                    llm=self.llm,
                    verbose=False,
                )

                task = Task(
                    description=prompt,
                    expected_output="Concise answer",
                    agent=agent,
                )

                crew = Crew(agents=[agent], tasks=[task], verbose=False)
                result = crew.kickoff()

                return str(result)
                
            except Exception as e:
                print(f"   ⚠️ LLM call attempt {attempt + 1} failed: {str(e)[:100]}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
                else:
                    print(f"   ❌ All LLM attempts failed, using fallback")
                    return "Manual investigation required - LLM service unavailable"

    def _aggressive_clean(self, text: str) -> str:
        """Aggressively remove LLM artifacts"""

        artifacts = [
            "I now can give",
            "FINAL ANSWER:",
            "Final Answer:",
            "Here is",
            "Here's",
            "The answer is:",
            "Answer:",
            "Step:",
            "Name:",
            "Explanation:",
            "Output:",
            "Current Task:",
            "My final response",
            "successfully",
        ]

        text_lower = text.lower()
        for artifact in artifacts:
            if artifact.lower() in text_lower:
                parts = text.split(artifact, 1)
                text = parts[-1] if len(parts) > 1 else parts[0]

        # Remove quotes, markdown, numbering
        text = re.sub(r'^["\'`\-\*\.]+', "", text.strip())
        text = re.sub(r'["\'`]+', "", text.strip())
        text = re.sub(r"\*\*", "", text)
        text = re.sub(r"```[a-z]*\n?", "", text)
        text = re.sub(r"^\d+\.\s*", "", text)
        text = re.sub(r"^:\s*", "", text)

        # Clean whitespace
        text = re.sub(r"\s+", " ", text)
        text = text.strip('" \n\r:.-')

        return text

    def _is_vague(self, text: str) -> bool:
        """Check if text is vague"""
        if not text or len(text) < 30:
            return True

        vague = ["n/a", "tbd", "complete the step", "document findings"]
        return any(phrase in text.lower() for phrase in vague)

    def export_to_excel(self, df: pd.DataFrame, rule_number: str) -> BytesIO:
        """Export to formatted Excel"""
        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Triaging_Template", index=False)
            workbook = writer.book
            worksheet = writer.sheets["Triaging_Template"]
            self._format_worksheet(worksheet, df)

        output.seek(0)
        return output

    def _format_worksheet(self, worksheet, df):
        """Apply Excel formatting"""
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        column_widths = {
            "A": 8,  # Step
            "B": 30,  # Name
            "C": 50,  # Explanation
            "D": 70,  # KQL Query (wider for better readability)
            "E": 40,  # KQL Explanation
            "F": 12,  # Execute
            "G": 30,  # Output
            "H": 35,  # Remarks
        }

        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        cell_alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = cell_alignment

                if row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color="F9F9F9", end_color="F9F9F9", fill_type="solid"
                    )

        worksheet.freeze_panes = "A2"


class EnhancedTemplateGenerator:
    def __init__(self):
        self.generator = ImprovedTemplateGenerator()
        self.template_columns = self.generator.template_columns

    def export_to_excel(self, df: pd.DataFrame, rule_number: str) -> BytesIO:
        return self.generator.export_to_excel(df, rule_number)
